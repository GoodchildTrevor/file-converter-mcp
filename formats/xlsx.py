"""XLSX exporter."""
from __future__ import annotations

import copy
import logging
import os
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.models import FileRef
from app.core.templates import TemplateRegistry
from storage.files import write_bytes
from storage.paths import new_export_folder, resolve_output_path, public_url

log = logging.getLogger(__name__)


def export_xlsx(
    rows: list[list[Any]],
    filename: str | None,
    title: str | None = None,
) -> FileRef:
    """Export *rows* as an Excel workbook and return a FileRef."""
    log.debug("Creating Excel workbook, rows=%d", len(rows))

    # Try to load a template if one is registered
    template_path = TemplateRegistry.get("xlsx")
    wb = None
    if template_path and os.path.exists(str(template_path)):
        try:
            wb = load_workbook(str(template_path))
            log.debug("XLSX template loaded from: %s", template_path)
        except Exception as exc:
            log.warning("Failed to load XLSX template: %s", exc)

    if wb is None:
        wb = Workbook()

    ws = wb.active

    if title:
        safe_title = "".join(c for c in title if c.isalnum() or c in (" ", "-", "_")).strip()[:31]
        ws.title = safe_title or "Sheet"

    if rows:
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = Font(bold=True)

        # Auto-fit column widths
        for c_idx in range(1, len(rows[0]) + 1):
            max_len = max(
                len(str(rows[r][c_idx - 1])) for r in range(len(rows)) if c_idx - 1 < len(rows[r])
            )
            ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 4, 80)

    # Save to bytes buffer then write via storage helper
    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    ext = "xlsx"
    folder = new_export_folder()
    filepath, fname = resolve_output_path(folder, filename or "", ext)
    with open(filepath, "wb") as f:
        f.write(buf.read())
    log.debug("write_xlsx → %s", filepath)
    return FileRef(path=filepath, name=fname, url=public_url(folder, fname))


def _create_excel(
    data: list[list[str]],
    filename: str,
    folder_path: str | None = None,
    title: str | None = None,
    template_path: str | None = None,
    template_obj=None,
) -> dict:
    """Legacy helper — kept for backward compatibility. Delegates to export_xlsx."""
    ref = export_xlsx(rows=data, filename=filename, title=title)
    return {"url": ref.url, "path": ref.path}
